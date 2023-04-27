from openpyxl import load_workbook
import zipfile
from pypdf import PdfReader
import xlrd
import csv
import time
import os.path
import requests
from selenium import webdriver
from selene import browser


# оформить в тест, добавить ассерты и использовать универсальный путь
def test_xlsx_file():
    xlsx_dir_path = os.path.dirname(os.path.abspath(__file__))
    xlsx_file_path = os.path.join(xlsx_dir_path, 'resources', 'file_example_XLSX_50.xlsx')
    workbook = load_workbook(xlsx_file_path)
    sheet1 = workbook.active
    value = sheet1.cell(row=4, column=2).value
    assert value == 'Philip'


# оформить в тест, добавить ассерты и использовать универсальный путь
def test_xls_file():
    xls_dir_path = os.path.dirname(os.path.abspath(__file__))
    xls_file_path = os.path.join(xls_dir_path, 'resources', 'file_example_XLS_10.xls')
    book = xlrd.open_workbook(xls_file_path)
    print(f'Количество листов {book.nsheets}')
    print(f'Имена листов {book.sheet_names()}')
    sheet = book.sheet_by_index(0)
    print(f'Количество столбцов {sheet.ncols}')
    print(f'Количество строк {sheet.nrows}')
    print(f'Пересечение строки 1 и столбца 1 = {sheet.cell_value(rowx=0, colx=1)}')
    # печать всех строк по очереди
    for rx in range(sheet.nrows):
        print(sheet.row(rx))

    assert book.nsheets == 1
    assert sheet.ncols == 8
    assert sheet.cell_value(rowx=0, colx=1) == 'First Name'


# оформить в тест, добавить ассерты и использовать универсальный путь
def test_pdf_file():
    pdf_dir_path = os.path.dirname(os.path.abspath(__file__))
    pdf_file_path = os.path.join(pdf_dir_path, 'resources', 'docs-pytest-org-en-latest.pdf')
    reader = PdfReader(pdf_file_path)
    number_of_pages = len(reader.pages)
    page = reader.pages[0]
    text = page.extract_text()
    print(page)
    print(number_of_pages)
    print(text)
    assert number_of_pages == 412


# оформить в тест, добавить ассерты и использовать универсальный путь
def test_csv_file():
    csv_dir_path = os.path.dirname(os.path.abspath(__file__))
    csv_file_path = os.path.join(csv_dir_path, 'resources', 'eggs.csv')
    with open(csv_file_path, 'w') as csvfile:
        csvwriter = csv.writer(csvfile, delimiter=',')
        csvwriter.writerow(['Anna', 'Pavel', 'Peter'])
        csvwriter.writerow(['Alex', 'Serj', 'Yana'])

    with open(csv_file_path) as csvfile:
        csvreader = csv.reader(csvfile)
        list = []
        for row in csvreader:
            list.append(row)
            print(row)
    assert list[0] == ['Anna', 'Pavel', 'Peter']


# сохранять и читать из tmp, использовать универсальный путь
def test_downloaded_file_size():
    url = 'https://selenium.dev/images/selenium_logo_square_green.png'
    downloaded_file_dir_path = os.path.dirname(os.path.abspath(__file__))
    downloaded_file_path = os.path.join(downloaded_file_dir_path, 'resources', 'selenium_logo.png')
    r = requests.get(url)
    with open(downloaded_file_path, 'wb') as file:
        file.write(r.content)
    size = os.path.getsize(downloaded_file_path)
    assert size == 30803


# оформить в тест, добавить ассерты и использовать универсальный путь к tmp
def test_download_file_browser():
    downloaded_file_dir_path = os.path.dirname(os.path.abspath(__file__))
    downloaded_file_path = os.path.join(downloaded_file_dir_path, 'tmp')

    options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": downloaded_file_path,
        "download.prompt_for_download": False
    }
    options.add_experimental_option("prefs", prefs)
    browser.config.driver_options = options

    browser.open('https://github.com/pytest-dev/pytest')
    browser.element('.d-none .Button-label').click()
    browser.element('[data-open-app="link"]').click()
    time.sleep(10)

    downloaded_file_name = os.path.join(downloaded_file_path, 'pytest-main.zip')
    downloaded_file_size = os.path.getsize(downloaded_file_name)
    assert downloaded_file_size > 0


def test_zip_file():
    zip_file_dir_path = os.path.dirname(os.path.abspath(__file__))
    zip_file_path = os.path.join(zip_file_dir_path, 'resources')
    zip_path = os.path.join(zip_file_dir_path, 'resources/archive.zip')
    file_zip = zipfile.ZipFile(zip_path, 'w')

    for folder, subfolders, files in os.walk(zip_file_path):
        for file in files:
            file_zip.write(os.path.join(folder, file), os.path.relpath(os.path.join(folder, file), 'resources'),
                           compress_type=zipfile.ZIP_DEFLATED)
    file_zip.close()

    file_lists = ['archive.zip', 'docs-pytest-org-en-latest.pdf', 'eggs.csv', 'file_example_XLSX_50.xlsx',
                 'file_example_XLS_10.xls', 'selenium_logo.png']

    lst = []
    with zipfile.ZipFile(zip_path, mode='a') as zf:
        for file in zf.namelist():
            lst.append(file)

    assert lst == file_lists
